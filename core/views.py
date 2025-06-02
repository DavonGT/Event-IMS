import json
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, JsonResponse
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Event, Organization, ExtensionActivity, College, Event

from .forms import EventForm, UploadFileForm, OrganizationForm, ExtensionActivityForm, CollegeForm
from django.views.decorators.http import require_GET
from openpyxl import load_workbook
from accounts.forms import UserProfileForm
from accounts.models import User
from django.urls import reverse
from django.db.models import F
from datetime import datetime, timedelta
from django.utils import timezone





@login_required
def events_dashboard(request):
    
    selected_org_id = request.GET.get('organization')  # Get org ID from query param
    organizations = Organization.objects.all()
    now = timezone.now()

    # Prepare events data for calendar
    events_json = [
        {   
            'name': event.name,
            'date': event.start_datetime.strftime('%Y-%m-%d'),
            'time': event.start_datetime.strftime('%H:%M'),
            'location': event.location,
            'description': event.description,
            'type': event.event_type,
            'organization': event.organization.name,
        }
        for event in Event.objects.all()
]
    
    # Get time range from request, default to week
    time_range = request.GET.get('time_range', 'week')
    
    # Calculate end date based on time range
    if time_range == 'week':
        end_date = now + timedelta(days=7)
    elif time_range == 'month':
        # Add one month by getting the first day of next month
        next_month = now.replace(day=1) + timedelta(days=32)
        end_date = next_month.replace(day=1)
    else:  # year
        end_date = now.replace(year=now.year + 1)

    # Get upcoming events based on time range
    base_upcoming_events = Event.objects.filter(
        end_datetime__gte=now,
        start_datetime__lte=end_date
    ).order_by('start_datetime')

    upcoming_events = base_upcoming_events[:10]
  # Show up to 10 upcoming events
    
    # Count all upcoming events for stats (keep this as weekly for the quick stats)
    week_end = now + timedelta(days=7)
    upcoming_events_count = Event.objects.filter(
        end_datetime__gte=now,
        start_datetime__lte=week_end
    ).count()
    
    # Calculate total events this year
    year_start = now.replace(month=1, day=1)
    total_events_this_year = Event.objects.filter(start_datetime__year=now.year).count()
    
    # Get total active organizations
    total_organizations = Organization.objects.count()

    # Calculate events per organization for pie chart
    org_events_data = []
    for org in organizations:
        event_count = Event.objects.filter(organization=org).count()
        if event_count > 0:  # Only include organizations with events
            org_events_data.append({
                'name': org.name,
                'acronym': org.acronym,
                'count': event_count
            })
    
    # Sort by count in descending order
    org_events_data.sort(key=lambda x: x['count'], reverse=True)

    # If an org is selected, filter events by that org
    selected_org = None
    if selected_org_id:
        try:
            selected_org = Organization.objects.get(id=selected_org_id)
            # Filter upcoming events based on selected organization
            upcoming_events = base_upcoming_events.filter(organization=selected_org)[:10]
            # Update total events for stats
            total_events_this_year = Event.objects.filter(
                organization=selected_org,
                start_datetime__year=now.year
            ).count()
            # Update upcoming events count (weekly stats)
            upcoming_events_count = Event.objects.filter(
                organization=selected_org,
                end_datetime__gte=now,
                start_datetime__lte=week_end
            ).count()
            
            # Prepare events data for calendar
            events_json = [
                {   
                    'name': event.name,
                    'date': event.start_datetime.strftime('%Y-%m-%d'),
                    'time': event.start_datetime.strftime('%H:%M'),
                    'location': event.location,
                    'description': event.description,
                    'type': event.event_type,
                    'organization': event.organization.name,
                }
                for event in Event.objects.all()
    ]


            # Add organization filter to calendar events
            events_json = [e for e in events_json if e['organization'] == selected_org.name]
        except Organization.DoesNotExist:
            selected_org_id = None
        
    # Get event counts by time period
    events_by_year = {}
    activities_by_year = {}
    current_year = now.year

    # Get data for the last 5 years
    for year in range(current_year - 4, current_year + 1):
        events_count = Event.objects.filter(
            start_datetime__year=year
        ).count()
        activities_count = ExtensionActivity.objects.filter(
            start_datetime__year=year
        ).count()
        
        events_by_year[year] = events_count
        activities_by_year[year] = activities_count

    # Get semester data for current year
    semesters_data = {
        '1st Semester': {
            'events': Event.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[6, 10]  # June to October
            ).count(),
            'activities': ExtensionActivity.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[6, 10]
            ).count()
        },
        '2nd Semester': {
            'events': Event.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[11, 12]  # November to December
            ).count() + Event.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[1, 3]  # January to March
            ).count(),
            'activities': ExtensionActivity.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[11, 12]
            ).count() + ExtensionActivity.objects.filter(
                start_datetime__year=current_year,
                start_datetime__month__range=[1, 3]
            ).count()
        }
    }

    

    context = {
        "user_role": str(request.user.role).title(),
        'organizations': organizations,
        'upcoming_events': upcoming_events,
        'upcoming_events_count': Event.objects.filter(end_datetime__gte=now, start_datetime__lte=end_date).count(),
        'time_range': time_range,
        'total_events_this_year': total_events_this_year,
        'total_organizations': total_organizations,
        'selected_org_id': selected_org_id,
        'selected_org': selected_org,
        'org_events_data': json.dumps(org_events_data),
        'events_by_year': json.dumps(events_by_year),
        'activities_by_year': json.dumps(activities_by_year),
        'semesters_data': json.dumps(semesters_data),
        'events_json': json.dumps(events_json, cls=DjangoJSONEncoder),
    }
    
    return render(request, 'core/dashboard.html', context)

    # user_role = str(request.user.role).title()
    #     return render(request, 'core/dashboard.html', {'user_role': user_role})








@login_required
def calendar(request):
    user = request.user
    events = Event.objects.all()
    organizations = Organization.objects.all()
    events_json = [
        {   
            'name': event.name,
            'date': event.start_datetime.strftime('%Y-%m-%d'),
            'time': event.start_datetime.strftime('%H:%M'),
            'location': event.location,
            'description': event.description,
            'type': event.event_type,
            'organization': event.organization.name,

        }
        for event in events
    ]

    if not request.user.is_authenticated:
        return redirect('login')

    # Role-based access control
    if request.user.role:
        return render(request, 'core/calendar.html', {
        "user": user,
        "events_json": json.dumps(events_json, cls=DjangoJSONEncoder),
        'user_role': str(request.user.role).title(),
        'organizations': organizations
    })
    else:
        return HttpResponseForbidden("You do not have permission to access this page.")  # Handle unknown roles

@login_required
def events_list(request, organization_id=None):
    organizations = Organization.objects.all()
    events = Event.objects.all()

    if organization_id:
        events = events.filter(organization_id=organization_id)
        selected_organization_id = organization_id
    else:
        selected_organization_id = None

    # Get unique event types and venues for filters
    event_types = list(events.order_by('event_type').values_list('event_type', flat=True).distinct())
    venues = list(events.order_by('location').values_list('location', flat=True).distinct())

    return render(request, 'core/events.html', {
        'events': events,
        'organizations': organizations,
        'selected_organization_id': selected_organization_id,
        'selected_organization': Organization.objects.get(id=selected_organization_id).name if selected_organization_id else None,
        'user_role': str(request.user.role).title(),
        'user_org':str(request.user.organization) if hasattr(request.user, 'organization') else None,
        'event_types': event_types,
        'venues': venues,
    })

@login_required
def activities_list(request, college_id=None):
    colleges = College.objects.all()
    activities = ExtensionActivity.objects.all()

    if college_id:
        activities = activities.filter(college_id=college_id)
        selected_college_id = college_id
    else:
        selected_college_id = None

    # Get unique event's venues for filters
    venues = list(activities.order_by('location').values_list('location', flat=True).distinct())
    print(colleges)
    return render(request, 'core/activities.html', {
        'activities': activities,
        'colleges': colleges,
        'selected_college_id': selected_college_id,
        'selected_college': College.objects.get(id=selected_college_id).name if selected_college_id else None,
        'user_role': str(request.user.role).title(),
        'venues': venues,
    })

# View to create a new event
@login_required
def add_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            event = form.save(commit=False)
            event.host = request.user
            # If organizer, ensure event.organization is set to user's org
            if hasattr(request.user, 'role') and request.user.role == 'organizer':
                event.organization = request.user.organization
            event.save()
            form.save_m2m()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('events_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string('core/partials/event_form.html', {'form': form}, request=request)
                return JsonResponse({'success': False, 'html': html})

    form = EventForm(user=request.user)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('core/partials/event_form.html', {'form': form}, request=request)
        return JsonResponse({'html': html})

    return render(request, 'core/events.html', {'form': form})

def delete_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Check if the user is the host of the event
    if request.user != event.host:
        messages.error(request, "You do not have permission to delete this event.")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
            return JsonResponse({'error': 'Permission Denied'}, status=403)
        return redirect(reverse('events_list', args=[id]))

    if request.method == 'POST':
        event.delete()
        messages.success(request, "Event deleted successfully.")
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
            return JsonResponse({'message': 'Event deleted successfully.', 'redirect_url': '/events/'})  # Redirect URL after delete

    # If the request is AJAX, return the delete modal HTML
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
        html = render(request, 'core/partials/delete_event_modal.html', {'event': event}).content.decode('utf-8')
        return JsonResponse({'html': html})

    return redirect('events_list')

def view_event(request, event_id):
    event = Event.objects.get(pk=event_id)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('core/partials/view_event_modal.html', {'event': event})
        return HttpResponse(html)
    
    return render(request, 'core/events.html', {'event': event})

def edit_event(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'message': 'Event updated successfully'})
            return redirect('events_list')
    
    else:
        form = EventForm(instance=event)
    return render(request, 'core/partials/edit_event_modal.html', {'form': form, 'event': event})

########################################################################################################################

# View to create a new activity
@login_required
def add_activity(request):
    user_role = str(request.user.role).title()
    # Get the selected college if passed, otherwise get the first one
    college_id = request.GET.get('college_id')
    if college_id:
        college = College.objects.get(id=college_id)
    else:
        college = College.objects.first()
        
    initial_data = {'college': college.id} if college else {}
    
    if request.method == 'POST':
        form = ExtensionActivityForm(request.POST, request.FILES)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.host = request.user
            activity.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('activities_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string('core/partials/activity_form.html', {
                    'form': form, 
                    'user_role': user_role,
                    'college': college
                }, request=request)
                return JsonResponse({'success': False, 'html': html})

    form = ExtensionActivityForm(initial=initial_data)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('core/partials/activity_form.html', {
            'form': form, 
            'user_role': user_role,
            'college': college
        }, request=request)
        return JsonResponse({'html': html})

    return render(request, 'core/activities.html', {
        'form': form,
        'user_role': user_role,
        'college': college
    })

def delete_activity(request, activity_id):
    activity = get_object_or_404(ExtensionActivity, id=activity_id)

    # Check if the user is the host of the activity
    if request.user != activity.host:
        messages.error(request, "You do not have permission to delete this activity.")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
            return JsonResponse({'error': 'Permission Denied'}, status=403)
        return redirect(reverse('activities_list'))

    if request.method == 'POST':
        activity.delete()
        messages.success(request, "Activity deleted successfully.")
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
            return JsonResponse({'message': 'Activity deleted successfully.', 'redirect_url': '/activities/'})  # Redirect URL after delete

    # If the request is AJAX, return the delete modal HTML
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request
        html = render(request, 'core/partials/delete_activity_modal.html', {'activity': activity}).content.decode('utf-8')
        return JsonResponse({'html': html})

    return redirect('activities_list')

def view_activity(request, activity_id):
    activity = get_object_or_404(ExtensionActivity, pk=activity_id)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('core/partials/view_activity_modal.html', {'activity': activity})
        return HttpResponse(html)
    
    return render(request, 'core/activities.html', {'activity': activity})

def edit_activity(request, activity_id):
    activity = get_object_or_404(ExtensionActivity, pk=activity_id)
    if request.method == 'POST':
        form = ExtensionActivityForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'message': 'Activity updated successfully'})
            return redirect('activities_list')
    
    else:
        form = ExtensionActivityForm(instance=activity)
    return render(request, 'core/partials/edit_activity_modal.html', {'form': form, 'activity': activity})

##############################################################################################################################

def add_organization(request):
    if request.method == 'POST':
        form = OrganizationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('events_list')  # Change to your preferred redirect
    else:
        form = OrganizationForm()

    return render(request, 'core/partials/add_org_modal.html', {'form': form})

def add_college(request):
    if request.method == 'POST':
        form = CollegeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('activities_list')  # Change to your preferred redirect
    else:
        form = CollegeForm()

    return render(request, 'core/partials/add_college_modal.html', {'form': form})

@require_GET
@login_required
def get_events_by_month(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return JsonResponse({'error': 'Missing parameters'}, status=400)

    events = Event.objects.filter(start_datetime__year=year, start_datetime__month=month)
    data = {}
    for event in events:
        org = event.organization.acronym
        if org not in data:
            data[org] = []
        data[org].append({
            'name': event.name,
            'date': event.start_datetime.strftime('%m.%d.%y'),
            'description': event.description,
            'location': event.location,
            'type': event.event_type,
        })

    return JsonResponse({'events_by_org': data})


def upload_file_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            organization_file = request.FILES.get('organization_file')
            event_file = request.FILES.get('event_file')

            if organization_file:
                wb = load_workbook(organization_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    Organization.objects.create(name=row[0], acronym=row[1], description=row[2])

            if event_file:
                wb = load_workbook(event_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    organization_acronym = row[0]
                    try: 
                        organization = Organization.objects.get(acronym=organization_acronym,)
                    except:
                        return redirect('events_list')
                    name = row[1].replace("'","").replace("`","")
                    description = row[2].replace("'","").replace("`","")
                    location = row[3].replace("'","").replace("`","")
                    
                    Event.objects.create(organization_id=organization.id, 
                                        name=name, description=description, 
                                        location=location, start_datetime=row[4], 
                                        end_datetime=row[5], event_type=row[6], 
                                        host=request.user)

            return redirect('events_list')
    else:
        form = UploadFileForm()

    return render(request, 'core/partials/upload_file.html', {
        'form': form,
        'user_role': str(request.user.role).title(),})


def profile(request):
    user = request.user
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect('profile')

    form = UserProfileForm(instance=user)
    return render(request, 'core/profile.html', {
        'user': user,
        'form':form,
        'user_role': str(user.role).title(),
    })

def edit_profile(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "User profile updated successfully.")
            return JsonResponse({'success': True, 'message': 'Profile updated successfully'})
    else:
        form = UserProfileForm(instance=user)
    return render(request, 'core/partials/edit_profile_modal.html', {'form': form, 'user': user})